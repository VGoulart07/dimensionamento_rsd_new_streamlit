import openpyxl
from openpyxl import Workbook
import math
import os

arquivo_excel = r"C:\Users\vinic\OneDrive - DSolar\DSOLAR\Advansol\Pré Vendas\Dimensionamento layout\dimensionamento_rsd.xlsx"

# Criar modelo de planilha se não existir
def criar_modelo_excel(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"

    ws["A1"] = "Potência do módulo (W)"
    ws["B1"] = 600
    ws["A2"] = "Potência do inversor (kW)"
    ws["B2"] = 75
    ws["A3"] = "Quantidade de strings conectadas"
    ws["B3"] = 16
    ws["A4"] = "Quantidade de módulos por string"
    ws["B4"] = 14
    ws["A5"] = "Modelo de RSD"
    ws["B5"] = "APT-MC-MR-T2"

    # Lista suspensa para RSD
    dv = openpyxl.worksheet.datavalidation.DataValidation(
        type="list",
        formula1='"APT-MC-R-T2,APT-MC-MR-T2,APT-MC-MRO"',
        allow_blank=False
    )
    ws.add_data_validation(dv)
    dv.add(ws["B5"])

    wb.save(path)
    print(f"📄 Modelo de planilha criado: {path}")

# Função para gerar a aba Saída
def gerar_saida_excel(path):
    wb = openpyxl.load_workbook(path)
    ws_in = wb["Inputs"]

    try:
        # Ler valores do Excel e converter para números
        potencia_modulo = float(str(ws_in["B1"].value).replace(",", "."))
        potencia_inversor = float(str(ws_in["B2"].value).replace(",", "."))
        qtd_strings = int(float(str(ws_in["B3"].value)))
        modulos_por_string = int(float(str(ws_in["B4"].value)))
        modelo_rsd = str(ws_in["B5"].value).strip()
    except Exception as e:
        print("❌ Erro na leitura dos inputs:", e)
        return

    # Debug: mostrar valores lidos
    print("=== Valores lidos da planilha ===")
    print(f"Potência do módulo: {potencia_modulo} W")
    print(f"Potência do inversor: {potencia_inversor} kW")
    print(f"Quantidade de strings: {qtd_strings}")
    print(f"Módulos por string: {modulos_por_string}")
    print(f"Modelo RSD: {modelo_rsd}")

    # Quantidade de RSD: 1 RSD para cada 2 módulos (arredondando para cima)
    rsd_por_string = math.ceil(modulos_por_string / 2)
    qtd_rsd_total = rsd_por_string * qtd_strings

    # Determinar controladora e quantidade de DCON
    if modelo_rsd == "APT-MC-R-T2":
        modelo_dcon = "APT-CB-DS"
        capacidade_dcon = 12
        qtd_dcon = math.ceil(qtd_strings / capacidade_dcon)
        nota = f"A controladora {modelo_dcon} suporta até {capacidade_dcon} strings, portanto foram necessárias {qtd_dcon} unidades."
    elif modelo_rsd in ["APT-MC-MR-T2", "APT-MC-MRO"]:
        if qtd_strings <= 4:
            modelo_dcon = "APT-CB-D-WIFI-S"
            capacidade_dcon = 4
            qtd_dcon = math.ceil(qtd_strings / capacidade_dcon)
            nota = f"A controladora {modelo_dcon} suporta até {capacidade_dcon} strings, portanto foram necessárias {qtd_dcon} unidades."
        elif qtd_strings <= 20:
            modelo_dcon = "APT-CB-D-WIFI-L"
            capacidade_dcon = 20
            qtd_dcon = math.ceil(qtd_strings / capacidade_dcon)
            nota = f"A controladora {modelo_dcon} suporta até {capacidade_dcon} strings, portanto foram necessárias {qtd_dcon} unidades."
        else:
            qtd_l = qtd_strings // 20
            resto = qtd_strings % 20
            qtd_s = math.ceil(resto / 4) if resto > 0 else 0
            qtd_dcon = qtd_l + qtd_s
            modelo_dcon = f"{qtd_l}x APT-CB-D-WIFI-L + {qtd_s}x APT-CB-D-WIFI-S"
            nota = f"Foi necessária a combinação: {modelo_dcon} para atender às {qtd_strings} strings."
    else:
        print("❌ Modelo de RSD inválido")
        return

    # Criar aba Saída
    if "Saída" in wb.sheetnames:
        ws_out = wb["Saída"]
        wb.remove(ws_out)
    ws_out = wb.create_sheet("Saída")

    ws_out.append(["Campo", "Valor"])
    ws_out.append(["Modelo RSD", modelo_rsd])
    ws_out.append(["Quantidade de RSD", qtd_rsd_total])
    ws_out.append(["Controladora de Dados", modelo_dcon])
    ws_out.append(["Quantidade de DCON", qtd_dcon])
    ws_out.append(["Nota explicativa", nota])

    wb.save(path)
    print(f"✅ Planilha atualizada com a aba Saída: {path}")
    print(f"Quantidade de RSD calculada: {qtd_rsd_total}")
    print(f"Controladora escolhida: {modelo_dcon} ({qtd_dcon} unidades)")

# Execução principal
if __name__ == "__main__":
    if not os.path.exists(arquivo_excel):
        criar_modelo_excel(arquivo_excel)
    gerar_saida_excel(arquivo_excel)
